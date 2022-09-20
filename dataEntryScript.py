from time import sleep
from dotenv import load_dotenv
import json
import os, openpyxl
import random, string
from hashlib import sha256
from datetime import date
from openpyxl import Workbook
from openpyxl import load_workbook
from sendMsgs import sendUsgNotif
from EmailSheet import sendMail, sendMailJson
from myLogger import myLogger
import myClasses

load_dotenv()
priceLoc = 'B'
minRow = '2'

myLog = myLogger("Automation-Logger", 10, "Event-Log" )

#Handles changeing of the date: being deprecated
def changeDate(newDate, sender):
    wb = None
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        makeTemplate(sender)
        wb = load_workbook(sheetP, data_only=True)
    ws = wb["Template"]
    day = newDate.split("/")[1]
    ws["N1"] = day
    pval = ws["N1"].value
    wb.save(sheetP)
    wb.close()

#Handles changing of the date with the json refactor
def changeDateJson(newDate, senderHash):
    user = getUser(senderHash)
    if "User does not exist" not in user:
        data = json.load(open('users.json', 'r'))
        day = newDate.split("/")[1]

        for x in data['users']:
            print(x['id'])
            if x['id'] == user['id']:
                print(x['id'])
                data['users'][data['users'].index(x)]['cycle_date'] = int(day)
                break

        json.dump(data, open('temp.json', 'w'), indent = 4)
        os.rename('temp.json', 'users.json')
    else:
        myLog.logInfo("User not found")

#makes template sheet in the workbook. Being deprecated
def makeTemplate(sender):
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    wb = Workbook()
    wsTemp = wb.create_sheet("Template")
    wsTemp['A1'] =  "Item / Location"
    wsTemp['B1'] = "Price"
    wsTemp['C1'] = "Date"
    wsTemp['D1'] = "Remaining: "
    wsTemp['F1'] = "Money Spent:  "
    wb.save(sheetP)
    wb.close()
    print(sender)

#makes template sheet in the workbook. JSON refactored
def makeTemplateJson(senderHash):
    user = getUser(senderHash)
    sheetP = os.getenv("sheetpath") + user['filename']
    wb = Workbook()
    wsTemp = wb.create_sheet("Template")
    wsTemp['A1'] =  "Item / Location"
    wsTemp['B1'] = "Price"
    wsTemp['C1'] = "Date"
    wsTemp['D1'] = "Remaining: "
    wsTemp['F1'] = "Money Spent:  "
    wb.save(sheetP)
    wb.close()
    print(user['username'])

#Removes duplicate sheets from the workbook
def removeDupes(sheetP):
    wb = load_workbook(sheetP, data_only=True)
    wsnL = []
    for sheetN in wb.sheetnames:
        if "Template Copy" in sheetN:
            wsnL.append(sheetN)
        if str(date.today()) in sheetN and len(sheetN) > len(str(date.today())):
            wsnL.append(sheetN)

    for sheetN in wsnL:
        wb.remove(wb[sheetN])

    """if "Template Copy" in wb.sheetnames:
        wb.remove(wb["Template Copy"])

    if str(date.today())in wb.sheetnames:
        wb.remove(wb[str(date.today()) + "1"])
    if str(date.today()) + "2" in wb.sheetnames:
        wb.remove(wb[str(date.today()) + "1"])"""
    wb.save(sheetP)
    wb.close()

#Manual Override the cycle turnover, being deprecated
def manualOverride(sender):
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    if os.path.exists(sheetP):
        handleTurn(sheetP)
        removeDupes(sheetP)
    else:
        makeTemplate(sender)
        handleTurn(sheetP)
        removeDupes(sheetP)

#Manual Override the cycle turnover, JSON Refactored
def manualOverJson(sendHash):
    sheetP = os.getenv("sheetpath") + getUser(sendHash)['filename']
    if "User not found" in sheetP and not os.path.exists(sheetP):
        pass
    else:
        handleTurn(sheetP)
        removeDupes(sheetP)

#Handles the turning over an individual sheet
def handleTurn(sheetP):
    wb = None
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        #makeTemplate(sender)
        wb = load_workbook(sheetP, data_only=True)

    if "Template Copy" in wb.sheetnames:
        wb.remove(wb["Template Copy"])

    if str(date.today()) + "1" in wb.sheetnames:
        wb.remove(wb[str(date.today()) + "1"])

    if "Template" in wb.sheetnames:
        ws = wb["Template"]
        temp = wb.copy_worksheet(ws)
        wb._add_sheet(temp, 0)
        #mylog.logInfo("Sheetnames Accessed: " +str(wb.sheetnames))
        ws = 0
        if "Template Copy" in wb.sheetnames:
            ws = wb["Template Copy"]
        if str(date.today()) in wb.sheetnames:
            wb.remove(ws)
            pass
        elif str(date.today()) not in wb.sheetnames:
            print("Adding sheet for today")
            ws.title = str(date.today())

        wb.save(sheetP)
        if "Template Copy" in wb.sheetnames:
            wb.remove(wb["Template Copy"])
        wb.save(sheetP)
        wb.close()
    else:
        wb.save(sheetP)
        wb.close()

#Runs automatically at 3am. Checks and turns over every sheet. Accounts for old structure and JSON refactor
def turnOver():
    fldrP = os.getenv("fldr")
    listSheets = os.listdir(fldrP)
    data = json.load(open('users.json', 'r'))['users']
    for fname in listSheets:
        filenme = os.path.join(fldrP, fname)
        if ".xlsx" in filenme and os.path.isfile(filenme):
            wb = load_workbook(filenme, data_only=True)
            wsT = wb["Template"]
            if wsT["N1"].value is None:
                for user in data:
                    if user['filename'] in filenme:
                        billDate = user["cycle_date"]
                        break
            else:
                billDate = wsT["N1"].value
            wb.save(filenme)
            month = int(wb.sheetnames[0].split("-")[1])
            wb.close()
            if int(date.today().day) == int(billDate):
                handleTurn(filenme)
                print("New Cycle, sheet has turned over")
                print(filenme)
                #mylog.logInfo("New Cycle, sheet has turned over")
                removeDupes(filenme)
            elif month == (int(date.today().month) - 1) and int(date.today().day) > int(billDate):
                handleTurn(filenme)
                print("New Cycle, sheet has turned over")
                print(filenme)
                #mylog.logInfo("New Cycle, sheet has turned over")
                removeDupes(filenme)
            else:
                pass
                #mylog.logInfo("Not Today")
        else:
            pass
            #mylog.logInfo("Not a file")"""

#Written, tested, and deprectaed before deployed
def reHash(sender):
    fldrP = os.getenv("fldr")
    listSheets = os.listdir(fldrP)
    for fname in listSheets:
        filenme = os.path.join(fldrP, fname)
        if str(sender[2:]) in filenme and os.path.isfile(filenme):
            tempNme = "Budgeting_Finances-"+ str(sha256(sender.encode('utf-8')).hexdigest()) + ".xlsx"
            newName = os.path.join(fldrP, tempNme)
            os.rename(filenme, newName)
            """wb = load_workbook(filenme, data_only=True)
            wsT = wb["Template"]
            billDate = wsT["N1"].value
            wb.save(filenme)
            month = int(wb.sheetnames[0].split("-")[1])
            wb.close()
            if int(date.today().day) == int(billDate):
                handleTurn(filenme)
                print("New Cycle, sheet has turned over")
                #mylog.logInfo("New Cycle, sheet has turned over")
                removeDupes(filenme)
            elif month == (int(date.today().month) - 1) and int(date.today().day) > int(billDate):
                handleTurn(filenme)
                print("New Cycle, sheet has turned over")
                #mylog.logInfo("New Cycle, sheet has turned over")
                removeDupes(filenme)
            else:
                pass
                #mylog.logInfo("Not Today")
        else:
            pass
            #mylog.logInfo("Not a file")"""

#Calculates the sum of the total spent. Being deprecated
def setupSum(sender):
    wb = None
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        wb = Workbook()
    ws = wb.active
    nums = []
    sum = 0.000
    for i in range(int(ws.max_row)):
        if ws['B' + str(i + 2)].value == None:
            break
        else:
            nums.append(float(ws['B' +  str(i + 2)].value))
    for x in range(len(nums)):
        sum= sum + float(nums[x])
    ws['G1'] = round(sum, 2)
    wsT = wb["Template"]
    cap = wsT["M1"].value
    ws['E1'] = round(float(cap) - sum, 2)
    wb.save(sheetP)
    wb.close()
    myLog.logInfo("Sum calculated")

#Handles making a purchase. Being deprecated
# Also runs a check to see if the cycle didn't turn over properly
def handleData(text, sender):
    #billDate = os.getenv("billDate")
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + sender + ".xlsx"
    vals = text.split('|')
    wb = None
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
        wsT = wb["Template"]
        billDate = wsT["N1"].value
        wb.close()
        if date.today().day == billDate:
            handleTurn(sheetP)
        wb = load_workbook(sheetP, data_only=True)

        #print(date.today().day)
    else:
        handleTurn(sheetP)
        wb = load_workbook(sheetP, data_only=True)

    ws = wb[wb.sheetnames[0]]
    if ws['G1'].value == None:
        wb.save(sheetP)
        wb.close()
        setupSum(sender)
    wb =  load_workbook(sheetP, data_only=True)
    ws = wb[wb.sheetnames[0]]
    if len(vals) > 1:
        price = vals[1][1:]
        itemLoc = 'A' + str(ws.max_row + 1)
        priceLoc= 'B' + str(ws.max_row + 1)
        dateLoc = 'C' + str(ws.max_row + 1)
        if ws['G1'].value is None:
            sum = 0.0
        else:
            sum = float(ws['G1'].value)
        sum+= float(price)
        ws['G1'] = sum
        ws['E1'] = round(float(wb["Template"]["M1"].value) - sum, 2)
        ws[itemLoc] = vals[0]
        ws[priceLoc] = float(price)
        ws[dateLoc] = str(date.today())
    wb.save(sheetP)
    wb.close()

#Generates an overview of purchases made this month. Being deprecated
def genOverview(sender):
    msg = ""
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    wb = load_workbook(sheetP, data_only=True)
    myLog.logInfo("Sheetnames Accessed: " + str(wb.sheetnames))
    ws = wb[wb.sheetnames[0]]
    msg = " \n Item | Price | Date "
    for i in range(ws.max_row):
        msg+="\n"
        if ws['A'  + str(i + 1)].value is not None:
            msg+= str(ws['A'  + str(i + 1)].value) + " | "
        else:
            msg+= "Item Missing | "
        if ws['B'  + str(i + 1)].value is not None:
            msg+=  str(ws['B'  + str(i + 1)].value) + " | "
        else:
            msg+= "Price Missing | "
        if ws['C'  + str(i + 1)].value is  None:
            msg+= "Date Missing "
        else:
            msg+=  str(ws['C'  + str(i + 1)].value)
    if ws["G1"].value is None:
        totalSpent = 0
    else:
        totalSpent = float(ws['G1'].value)
    cap = wb["Template"]["M1"].value
    if totalSpent is None:
        totalSpent =  0
    totalRemaining = round(float(cap) - totalSpent, 2) #TODO read budget cap from sheet
    wb.save(sheetP)
    wb.close()
    msg+= "\nTotal Spent ($USD): " + str(round(totalSpent, 2))
    msg+= "\nTotal Remaining ($USD): " + str(totalRemaining)
    return msg

#Formats a msg containing the last purchase, total spent, total remaining, being deprecated
def formatMsg(sender):
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(sender) + ".xlsx"
    wb = None
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    mr = ws.max_row
    pLoc = 'B' + str(mr)
    iLoc = 'A' + str(mr)
    item = ws[iLoc].value
    cost = ws[pLoc].value
    if ws['G1'].value == None:
        ws['G1'] = 0.0
        setupSum(sender)
    totalSpent = float(ws['G1'].value)
    if totalSpent is None:
        totalSpent =  0
    totalRemaining = round(float(wb["Template"]["M1"].value) - totalSpent, 2)
    wb.save(sheetP)
    wb.close()

    if item is None or cost is None:
        msg =  "Empty Sheet"
    else:
        msg = "\nItem: " + str(item)
        msg+= "\nPrice ($USD): " + str(round(cost, 2))
        msg+= "\nTotal Spent ($USD): " + str(round(totalSpent, 2))
        msg+= "\nTotal Remaining ($USD): " + str(totalRemaining)
    return msg

#Formats a msg containing the last purchase, total spent, total remaining. JSON refactored
def formatMsgJson(senderHash):
    user = getUser(senderHash)
    sheetP = os.getenv("sheetpath") + user['filename']

    wb = None
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        wb = Workbook()
    ws = wb[wb.sheetnames[0]]
    mr = ws.max_row
    pLoc = 'B' + str(mr)
    iLoc = 'A' + str(mr)
    item = ws[iLoc].value
    cost = ws[pLoc].value
    if ws['G1'].value == None:
        ws['G1'] = 0.0
        setupSumJson(senderHash)
    totalSpent = float(ws['G1'].value)
    if totalSpent is None:
        totalSpent =  0
    totalRemaining = round(float(user['budget']) - totalSpent, 2)
    wb.save(sheetP)
    wb.close()

    if item is None or cost is None:
        msg =  "Empty Sheet"
    else:
        msg = "\nItem: " + str(item)
        msg+= "\nPrice ($USD): " + str(round(cost, 2))
        msg+= "\nTotal Spent ($USD): " + str(round(totalSpent, 2))
        msg+= "\nTotal Remaining ($USD): " + str(totalRemaining)
    return msg

#Handles account creation through text. Being deprecated
def initNewAccount(sender, billDate, budgCap):
    sheetP = os.getenv("sheetpath")
    makeTemplate(sender)
    sheetP = sheetP + str(sender) + ".xlsx"
    handleTurn(sheetP)
    wb = load_workbook(sheetP, data_only=True)
    ws = wb["Template"]
    ws["M1"] = budgCap
    ws["N1"] = billDate
    wb.save(sheetP)
    wb.close()
    return ("Sheet Made: use format \nItem | Price\n to add an item")

#Generates an overview of purchases made this month. JSON Refactored
def genOverviewJson(senderHash):
    msg = ""
    user = getUser(senderHash)
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + user['filename']
    wb = load_workbook(sheetP, data_only=True)
    myLog.logInfo("Sheetnames Accessed: " + str(wb.sheetnames))
    ws = wb[wb.sheetnames[0]]
    msg = " \n Item | Price | Date "
    for i in range(ws.max_row):
        msg+="\n"
        if ws['A'  + str(i + 1)].value is not None:
            msg+= str(ws['A'  + str(i + 1)].value) + " | "
        else:
            msg+= "Item Missing | "
        if ws['B'  + str(i + 1)].value is not None:
            msg+=  str(ws['B'  + str(i + 1)].value) + " | "
        else:
            msg+= "Price Missing | "
        if ws['C'  + str(i + 1)].value is  None:
            msg+= "Date Missing "
        else:
            msg+=  str(ws['C'  + str(i + 1)].value)
    if ws["G1"].value is None:
        totalSpent = 0
    else:
        totalSpent = float(ws['G1'].value)
    cap = float(user['budget'])
    if totalSpent is None:
        totalSpent =  0
    totalRemaining = round(float(cap) - totalSpent, 2) #TODO read budget cap from sheet
    wb.save(sheetP)
    wb.close()
    msg+= "\nTotal Spent ($USD): " + str(round(totalSpent, 2))
    msg+= "\nTotal Remaining ($USD): " + str(totalRemaining)
    return msg

#Calculates the sum of the total spent in the month/ JSON Refactored
def setupSumJson(sendHash):
    wb = None
    user = getUser(sendHash)
    sheetP = os.getenv("sheetpath") + user['filename']
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
    else:
        wb = Workbook()
    ws = wb.active
    nums = []
    sum = 0.000
    for i in range(int(ws.max_row)):
        if ws['B' + str(i + 2)].value == None:
            break
        else:
            nums.append(float(ws['B' +  str(i + 2)].value))
    for x in range(len(nums)):
        sum= sum + float(nums[x])
    ws['G1'] = round(sum, 2)
    cap = user['budget']
    ws['E1'] = round(float(cap) - sum, 2)
    wb.save(sheetP)
    wb.close()
    myLog.logInfo("Sum calculated")

#Handles making a purchase. JSON Refactored
# Also runs a check to see if the cycle didn't turn over properly
def handleDataFromJson(text, sender):
    user = getUser(sender)
    sheetP = os.getenv("sheetpath") + user['filename']
    vals = text.split('|')
    wb = None
    if os.path.exists(sheetP):
        wb = load_workbook(sheetP, data_only=True)
        wb.close()
        if date.today().day == user['cycle_date']:
            handleTurn(sheetP)
        wb = load_workbook(sheetP, data_only=True)

    else:
        handleTurn(sheetP)
        wb = load_workbook(sheetP, data_only=True)

    ws = wb[wb.sheetnames[0]]
    if ws['G1'].value == None:
        wb.save(sheetP)
        wb.close()
        setupSumJson(sender)
    wb =  load_workbook(sheetP, data_only=True)
    ws = wb[wb.sheetnames[0]]
    if len(vals) > 1:
        price = vals[1][1:]
        itemLoc = 'A' + str(ws.max_row + 1)
        priceLoc= 'B' + str(ws.max_row + 1)
        dateLoc = 'C' + str(ws.max_row + 1)
        if ws['G1'].value is None:
            sum = 0.0
        else:
            sum = float(ws['G1'].value)
        sum+= float(price)
        ws['G1'] = sum
        ws['E1'] = round(float(float(user['budget']) - sum), 2)
        ws[itemLoc] = vals[0]
        ws[priceLoc] = float(price)
        ws[dateLoc] = str(date.today())
    wb.save(sheetP)
    wb.close()
    del user

#makes template sheet in the workbook. JSON Refactored
def makeTempJson(sheetP):
    wb = Workbook()
    wsTemp = wb.create_sheet("Template")
    wsTemp['A1'] =  "Item / Location"
    wsTemp['B1'] = "Price"
    wsTemp['C1'] = "Date"
    wsTemp['D1'] = "Remaining: "
    wsTemp['F1'] = "Money Spent:  "
    wb.save(sheetP)
    wb.close()

#Handles account creation through text. JSON Refactored
def initJsonAccount(sender, un, billDate, budg):
    fname = makeJsonData(sender, sha256(sender.encode('utf-8')).hexdigest(), un, budg, billDate)
    sheetP = os.getenv("sheetpath") + fname
    makeTempJson(sheetP)
    handleTurn(sheetP)

#Adds user to the json data structure. Useful for backend management and helps with privacy a bit more
def makeJsonData(sender, senderHash, un, budg, cycle):
    listIds = []
    data = json.load(open('users.json', 'r'))
    if ".xlsx" not in un:
        un+= ".xlsx"

    for x in data['users']:
        listIds.append(x['id'])
    if senderHash in listIds:
        return data['users'][listIds.index(senderHash)]['filename']

    else:
        word = sender.join(random.choice(string.ascii_letters) for i in range(26))
        fn = sha256(word.encode('utf-8')).hexdigest() + ".xlsx"
        newUser = myClasses.User(senderHash, fn, un, int(budg), int(cycle))
        data['users'].append(newUser.__dict__)
        json.dump(data, open('temp.json', 'w'), indent = 4)
        os.rename('temp.json', 'users.json')
        print("writing json")
        return fn

#Translates a preexisting account to use the new json structure
def jsonIfy(sender, senderHash, un):
    p = os.getenv("sheetpath")
    sheetP = None
    if os.path.exists(p + sender + ".xlsx"):
        sheetP = p + sender + ".xlsx"
    else:
        sheetP = p + "Budgeting-Finances" + sender + ".xlsx"
    #print(sheetP)
    wb = load_workbook(sheetP, data_only=True)
    wst = wb["Template"]
    budg = wst["M1"].value
    cycleD = wst["N1"].value
    fn = os.getenv("sheetpath") + makeJsonData(sender, senderHash, un, budg, cycleD)
    wst['M1'] = None
    wst['N1'] = None
    wb.save(sheetP)
    wb.close()
    os.rename(sheetP, fn)
    print(genOverviewJson(senderHash))


#Gets user based off hash id
def getUser(sendHash):
    data = json.load(open('users.json', 'r'))
    y = None
    for x in data['users']:
        if x['id'] == sendHash:
            y = x
            break
    if y is None:
        y = "User does not exist!"

    return y

def checkAuthUser(sendHash):
    data = json.load(open('users.json', 'r'))
    val = False
    for x in data['users']:
        if x['id'] == sendHash:
            val = True

    return val

#Emails sheet, being deprecated
def sendSheet(addr, filenme, sender):
    sheetP = os.getenv("sheetpath")
    sheetP = sheetP + str(filenme) + ".xlsx"
    if os.path.exists(sheetP):
        sendMail(addr, sheetP, sender)
        myLog.logInfo("Email Sent")
        return("Your sheet was sent to: " + addr)
    else:
        myLog.logInfo("File not found, please generate a spreadsheet to email it to someone")
        return("File not found, please generate a spreadsheet to email it to someone")

#Emails sheet, JSON Refactored
def sendSheetJson(addr, senderHash):
    user = getUser(senderHash)
    sheetP = os.getenv("sheetpath") + user['filename']
    if os.path.exists(sheetP):
        sendMailJson(addr, sheetP, user)
        myLog.logInfo("Email Sent")
        return("Your sheet was sent to: " + addr)
    else:
        myLog.logInfo("File not found, please generate a spreadsheet to email it to someone")
        return("File not found, please generate a spreadsheet to email it to someone")