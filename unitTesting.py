from hashlib import sha256, sha512
import dataEntryScript
from dotenv import load_dotenv
import os
from myLogger import myLogger
import EmailSheet
import openpyxl as opxl
import threading
import sendMsgs
import argparse
from time import sleep

load_dotenv()

parser = argparse.ArgumentParser(description="Handles excluding specific testting suites. So I can test without proper networking setup.")
parser.add_argument('--email', help='Disables email tests for testing in different networks environment', default = 1, type=int)
parser.add_argument('--msgSend', help='Disables the send message testing for different development environments', default=True, type= bool)

args = parser.parse_args()
sheetP = os.getenv("sheetpath")
sheetPOne = sheetP + "1234567890"+ ".xlsx"
testingLog = myLogger("Testing-Logger", 10, "Testing-Log")

#TESTING INIT NEW ACCOUNT AND HANDLE DATA

testingLog.logDebug("\n\n\tTESTING INIT AND DATA HANDLING\t\n\n")
dataEntryScript.initNewAccount("1234567890", 17, 6969)
dataEntryScript.handleData("Mike's Apple | 25", "1234567890")
dataEntryScript.handleData("Taza | 50", "1234567890")
dataEntryScript.handleData("Chipotle | 25", "1234567890")
dataEntryScript.handleData("Baja Blasted Vodka| 60", "1234567890")
dataEntryScript.handleData("Fairlife IceCream | 4.59", "1234567890")
testingLog.logInfo(dataEntryScript.genOverview("1234567890"))
testingLog.logInfo(dataEntryScript.formatMsg("1234567890"))

#TESTING JSON REFACTOR
dataEntryScript.initJsonAccount("+10987654321", "Unit-Testing.xlsx", 17, 6969)
#print(dataEntryScript.getUser(sha256(b"+10987654321").hexdigest()))
dataEntryScript.handleDataFromJson("Mike's Apple | 25", sha256(b"+10987654321").hexdigest())
dataEntryScript.handleDataFromJson("Taza | 50", sha256(b"+10987654321").hexdigest())
dataEntryScript.handleDataFromJson("Chipotle | 25", sha256(b"+10987654321").hexdigest())
dataEntryScript.handleDataFromJson("Baja Blasted Vodka| 60", sha256(b"+10987654321").hexdigest())
dataEntryScript.handleDataFromJson("Fairlife IceCream | 4.59", sha256(b"+10987654321").hexdigest())
sleep(4)
dataEntryScript.changeDateJson("09/21/2002", sha256(b"+10987654321").hexdigest())
dataEntryScript.removeDupes(sheetP + dataEntryScript.getUser(sha256(b"+10987654321").hexdigest())['filename'])
dataEntryScript.manualOverJson(sha256(b"+10987654321").hexdigest())



print(args.email)
#TESTING EMAILING
if args.email == 1:
    testingLog.logDebug("\n\n TESTING EMAIL FUNCTIONALITY \n\n")
    testingLog.logInfo(dataEntryScript.sendSheet("williamryan978@icloud.com", "1234567890", "1234567890"))
    EmailSheet.sendMail("williamryan978@icloud.com", sheetPOne, "1234567890")
else:
    testingLog.logDebug("\n\n SKPPING EMAIL TESTS \n\n")
#TESTING SETUP SUM
testingLog.logDebug("\n\n\tTESTING SUM CORRECTION\t\n\n")
dataEntryScript.setupSum("1234567890")
testingLog.logInfo(dataEntryScript.genOverview("1234567890"))
wb = opxl.load_workbook(sheetPOne, data_only=True)
ws = wb[wb.sheetnames[0]]
ws['G1'] = 5555
wb.save(sheetPOne)
wb.close()

testingLog.logInfo(dataEntryScript.genOverview("1234567890"))
dataEntryScript.setupSum("1234567890")
testingLog.logInfo(dataEntryScript.genOverview("1234567890"))

#TESTING HANDLE TURN, TURNOVER, CHANGEDATE, AND MAN OVERRIDE
testingLog.logDebug("\t\n\nTESTING CYCLE TURNOVERS, auto and manual\t\n\n")
dataEntryScript.turnOver()
dataEntryScript.changeDate("1/12/11", "1234567890")
dataEntryScript.manualOverride("1234567890")
testingLog.logInfo(dataEntryScript.genOverview("1234567890"))
testingLog.logInfo(dataEntryScript.formatMsg("1234567890"))
dataEntryScript.handleTurn(sheetPOne)

#TESTING LOGGING ON MORE A DIFFERENT THREAD
testingLog.logDebug("\n\n\t TESTING LOGGER ON DIFFERENT THREAD, MAY BREAK \t\n\n")
def tag():
    words = ["Hi", " world", "new", "apples", "are", "nice"]
    for x in range(len(words)):
        testingLog.logWarn("\n\n" + words[x] + "\n\n")
        sleep(1)

testingThread = threading.Thread(target=tag)
testingThread.start()
testingThread.join()

testingLog.logInfo("\n\n\tTesting two threads at once\t\n\n")
#TESTING LOGGING ON TWO THREADS AT ONCE
testingThread2 = threading.Thread(target=tag)
testingThread2.start()
sleep(2)
testingLog.logWarn("Two threads simultaneously")
testingThread2.join()
#TESTING SENDUSGNOTIF
#sendMsgs.sendUsgNotif("\n\n\tUnit Testing \n\n")


#TESTTING HASH DEVELOPMENT
testingLog.logWarn("\n\tHASHING TESTS: EXPECT SLOWDOWN\n")
print(sha512(b"+11234567890").hexdigest())
print(sha256(b"+11234567890").hexdigest())
print(hash("+11234567890"))
print(hash("+11234567890"))
dataEntryScript.jsonIfy("1234567890", sha256(b"+11234567890").hexdigest(), "Unit Tester")

#dataEntryScript.makeJsonData(sha256(b"+17817750100").hexdigest(), "protect the turtles", 1700, 17)